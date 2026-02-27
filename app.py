"""
app.py  –  메디게이트 구인 트렌드 Streamlit 대시보드

실행
----
    streamlit run app.py

기능
----
- 사이드바: 지역 / 진료과 필터 (드롭다운)
- 막대그래프: 월별 구인건수 (Plotly)
- 막대 클릭 → 팝업 다이얼로그: 해당 월 병원 목록 표시
- 급여 현황: 지역별 / 진료과별 평균 Net 월급 수평 막대 그래프
"""

import os
import re as _re
from datetime import datetime as _dt

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from sqlalchemy import create_engine, text

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "(마봉협)구인구직정리.xlsx",
)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 페이지 설정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.set_page_config(
    page_title="메디게이트 구인 트렌드",
    page_icon="🏥",
    layout="wide",
)

DB_URL = "postgresql+psycopg2://postgres:postgres@localhost:5432/medigate"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DB 엔진 (앱 생명주기 동안 1회만 생성)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@st.cache_resource
def get_engine():
    return create_engine(DB_URL)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 마취통증의학과 전용 — 엑셀 과거자료 로드 (1시간 캐싱)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@st.cache_data(ttl=3600)
def load_excel_machwi() -> pd.DataFrame:
    """엑셀 일자리분석 시트에서 월별 공고수 & 평균 Net 월급 추출.
    구조: 4열 1그룹 (blank | 지역 | 병원명 | Pay)  ×  35개월
    Pay 단위: 2.3 → 2,300만원 (× 1000)
    """
    if not os.path.exists(EXCEL_PATH):
        return pd.DataFrame()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["일자리분석"]
    except Exception:
        return pd.DataFrame()

    records = []
    col = 1
    while col <= ws.max_column:
        # 그룹 날짜: row 8, 각 그룹 첫 번째 열
        date_val = ws.cell(8, col).value
        if date_val is None:
            col += 4
            continue

        if isinstance(date_val, _dt):
            month_str = date_val.strftime("%Y-%m")
        elif isinstance(date_val, str):
            m = _re.match(r"(\d{4})년(\d{1,2})월", date_val)
            month_str = f"{m.group(1)}-{int(m.group(2)):02d}" if m else None
        else:
            month_str = None

        if not month_str:
            col += 4
            continue

        # Pay 수집: col+3 (4번째 열)
        pays = []
        for row in range(10, ws.max_row + 1):
            v = ws.cell(row, col + 3).value
            if v is None:
                continue
            try:
                f = float(v)
                if 0.5 <= f <= 10.0:          # 500 ~ 10,000만원 유효 범위
                    pays.append(f * 1000)
            except (TypeError, ValueError):
                continue

        if pays:
            records.append({
                "등록월":    month_str,
                "공고수":    len(pays),
                "평균Net월급": round(sum(pays) / len(pays)),
                "출처":      "엑셀(과거)",
            })
        col += 4

    df = pd.DataFrame(records)
    if not df.empty:
        df = df.sort_values("등록월").drop_duplicates("등록월", keep="first")
    return df


@st.cache_data(ttl=60)
def load_db_machwi() -> pd.DataFrame:
    """DB에서 마취통증의학과 월별 공고수 & 평균 Net 월급 추출.
    - 공고수: 전체 마취통증의학과 공고 (급여 여부 무관)
    - 평균Net월급: salary_type=net, salary_unit=monthly, >500만원 한정
    """
    sql_cnt = text("""
        SELECT
            LEFT(rp.register_date, 7)  AS reg_month,
            COUNT(DISTINCT rp.id)      AS post_cnt
        FROM  recruit_posts rp
        JOIN  recruit_post_specialties rps ON rps.post_id = rp.id
        WHERE rps.specialty LIKE '%마취%'
          AND rp.register_date IS NOT NULL
          AND rp.register_date <> ''
        GROUP BY LEFT(rp.register_date, 7)
        ORDER BY reg_month
    """)
    sql_sal = text("""
        SELECT
            LEFT(rp.register_date, 7)                                  AS reg_month,
            ROUND(AVG((rp.salary_net_min + rp.salary_net_max) / 2.0))  AS avg_net
        FROM  recruit_posts rp
        JOIN  recruit_post_specialties rps ON rps.post_id = rp.id
        WHERE rps.specialty LIKE '%마취%'
          AND rp.register_date IS NOT NULL
          AND rp.register_date <> ''
          AND rp.salary_type  = 'net'
          AND rp.salary_unit  = 'monthly'
          AND rp.salary_net_min > 500
          AND rp.salary_net_max > 500
        GROUP BY LEFT(rp.register_date, 7)
        ORDER BY reg_month
    """)
    try:
        with get_engine().connect() as conn:
            df_cnt = pd.read_sql(sql_cnt, conn)
            df_sal = pd.read_sql(sql_sal, conn)
        df = df_cnt.merge(df_sal, on="reg_month", how="left")
        df = df.rename(columns={"reg_month": "등록월", "post_cnt": "공고수", "avg_net": "평균Net월급"})
        df["출처"] = "DB(크롤링)"
        return df
    except Exception as e:
        st.error(f"마취통증 DB 조회 오류: {e}")
        return pd.DataFrame()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 집계 데이터 로드 (60초 캐싱)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@st.cache_data(ttl=60)
def load_salary_monthly(region: str, specialty: str) -> pd.DataFrame:
    """월별 평균 Net 월급 집계 (메인 차트용, 봉직의 한정)."""
    conditions = [
        "rp.salary_net_min IS NOT NULL",
        "rp.salary_net_max IS NOT NULL",
        "rp.register_date IS NOT NULL",
        "rp.register_date <> ''",
        "rp.employment_type = '봉직의'",
        "(rp.salary_net_min + rp.salary_net_max) / 2.0 > 1000",
    ]
    params: dict = {}
    need_join = specialty != "전체"

    if region != "전체":
        if len(region) > 2:  # 시도+시군 조합 (예: 경기수원, 경북포항)
            sido = region[:2]
            city = region[2:]
            conditions.append("rp.region_sido = :region_sido")
            conditions.append("rp.region LIKE :region_like")
            params["region_sido"] = sido
            params["region_like"] = f"{sido} {city}%"
        else:
            conditions.append("rp.region_sido = :region")
            params["region"] = region
    if specialty != "전체":
        conditions.append("rps.specialty = :specialty")
        params["specialty"] = specialty

    where = " AND ".join(conditions)
    join  = "JOIN recruit_post_specialties rps ON rps.post_id = rp.id" if need_join else ""

    sql = text(f"""
        WITH base AS (
            SELECT
                LEFT(rp.register_date, 7)                      AS reg_month,
                (rp.salary_net_min + rp.salary_net_max) / 2.0 AS salary_mid
            FROM recruit_posts rp
            {join}
            WHERE {where}
        ),
        stats AS (
            SELECT
                reg_month,
                COUNT(*)                                                  AS cnt,
                PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY salary_mid) AS q1,
                PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY salary_mid) AS q3,
                PERCENTILE_CONT(0.5)  WITHIN GROUP (ORDER BY salary_mid) AS median_val
            FROM base
            GROUP BY reg_month
        ),
        filtered AS (
            SELECT b.reg_month, b.salary_mid, s.cnt, s.median_val
            FROM base b
            JOIN stats s ON s.reg_month = b.reg_month
            WHERE s.cnt < 15
               OR b.salary_mid BETWEEN s.q1 - 1.5 * (s.q3 - s.q1)
                                   AND s.q3 + 1.5 * (s.q3 - s.q1)
        )
        SELECT
            reg_month,
            ROUND(CASE WHEN MAX(cnt) >= 15 THEN AVG(salary_mid)
                       ELSE MAX(median_val) END) AS avg_net,
            MAX(cnt)                             AS cnt
        FROM filtered
        GROUP BY reg_month
        ORDER BY reg_month
    """)
    try:
        with get_engine().connect() as conn:
            df = pd.read_sql(sql, conn, params=params)
        return df.rename(columns={"reg_month": "등록월", "avg_net": "평균Net월급", "cnt": "공고수"})
    except Exception as e:
        st.error(f"급여 월별 조회 오류: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=60)
def load_salary_ranking(region: str, specialty: str) -> tuple:
    """지역별 / 진료과별 전체 평균 순위 (보조 테이블용, 봉직의 한정)."""
    params: dict = {}
    need_join = specialty != "전체"
    conditions_base = [
        "rp.salary_net_min IS NOT NULL",
        "rp.salary_net_max IS NOT NULL",
        "rp.employment_type = '봉직의'",
        "(rp.salary_net_min + rp.salary_net_max) / 2.0 > 1000",
    ]
    if region != "전체":
        if len(region) > 2:  # 시도+시군 조합 (예: 경기수원, 경북포항)
            sido = region[:2]
            city = region[2:]
            conditions_base.append("rp.region_sido = :region_sido")
            conditions_base.append("rp.region LIKE :region_like")
            params["region_sido"] = sido
            params["region_like"] = f"{sido} {city}%"
        else:
            conditions_base.append("rp.region_sido = :region")
            params["region"] = region
    if specialty != "전체":
        conditions_base.append("rps.specialty = :specialty")
        params["specialty"] = specialty

    join  = "JOIN recruit_post_specialties rps ON rps.post_id = rp.id" if need_join else ""
    where = " AND ".join(conditions_base)

    # 지역별 순위 (시도 단위)
    sql_r = text(f"""
        WITH base AS (
            SELECT
                rp.region_sido                                 AS region,
                (rp.salary_net_min + rp.salary_net_max) / 2.0 AS salary_mid
            FROM recruit_posts rp {join}
            WHERE {where}
              AND rp.region_sido IS NOT NULL AND rp.region_sido <> ''
        ),
        stats AS (
            SELECT
                region,
                COUNT(*)                                                  AS cnt,
                PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY salary_mid) AS q1,
                PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY salary_mid) AS q3,
                PERCENTILE_CONT(0.5)  WITHIN GROUP (ORDER BY salary_mid) AS median_val
            FROM base
            GROUP BY region
        ),
        filtered AS (
            SELECT b.region, b.salary_mid, s.cnt, s.median_val
            FROM base b
            JOIN stats s ON s.region = b.region
            WHERE s.cnt < 15
               OR b.salary_mid BETWEEN s.q1 - 1.5 * (s.q3 - s.q1)
                                   AND s.q3 + 1.5 * (s.q3 - s.q1)
        )
        SELECT
            region,
            ROUND(CASE WHEN MAX(cnt) >= 15 THEN AVG(salary_mid)
                       ELSE MAX(median_val) END) AS avg_net,
            MAX(cnt)                             AS cnt
        FROM filtered
        GROUP BY region
        ORDER BY avg_net DESC
    """)
    # 진료과별 순위
    join_s  = "JOIN recruit_post_specialties rps ON rps.post_id = rp.id"
    cond_s  = [c for c in conditions_base if "rps.specialty" not in c]
    where_s = " AND ".join(cond_s)
    sql_s = text(f"""
        WITH base AS (
            SELECT
                rps.specialty                                  AS specialty,
                (rp.salary_net_min + rp.salary_net_max) / 2.0 AS salary_mid
            FROM recruit_posts rp {join_s}
            WHERE {where_s}
        ),
        stats AS (
            SELECT
                specialty,
                COUNT(*)                                                  AS cnt,
                PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY salary_mid) AS q1,
                PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY salary_mid) AS q3,
                PERCENTILE_CONT(0.5)  WITHIN GROUP (ORDER BY salary_mid) AS median_val
            FROM base
            GROUP BY specialty
        ),
        filtered AS (
            SELECT b.specialty, b.salary_mid, s.cnt, s.median_val
            FROM base b
            JOIN stats s ON s.specialty = b.specialty
            WHERE s.cnt < 15
               OR b.salary_mid BETWEEN s.q1 - 1.5 * (s.q3 - s.q1)
                                   AND s.q3 + 1.5 * (s.q3 - s.q1)
        )
        SELECT
            specialty,
            ROUND(CASE WHEN MAX(cnt) >= 15 THEN AVG(salary_mid)
                       ELSE MAX(median_val) END) AS avg_net,
            MAX(cnt)                             AS cnt
        FROM filtered
        GROUP BY specialty
        HAVING MAX(cnt) >= 5
        ORDER BY avg_net DESC
    """)
    try:
        with get_engine().connect() as conn:
            df_r = pd.read_sql(sql_r, conn, params=params).rename(
                columns={"region": "지역", "avg_net": "평균Net월급", "cnt": "공고수"})
            df_s = pd.read_sql(sql_s, conn, params={k: v for k, v in params.items()
                                                    if k != "specialty"}).rename(
                columns={"specialty": "진료과", "avg_net": "평균Net월급", "cnt": "공고수"})
        return df_r, df_s
    except Exception as e:
        st.error(f"순위 조회 오류: {e}")
        return pd.DataFrame(), pd.DataFrame()


@st.cache_data(ttl=60)
def load_aggregated() -> pd.DataFrame:
    """(region, specialty, employment_type, reg_month, post_count) 집계 테이블 반환."""
    try:
        with get_engine().connect() as conn:
            return pd.read_sql(text("""
                SELECT
                    rp.region_sido            AS region,
                    rps.specialty             AS specialty,
                    rp.employment_type        AS employment_type,
                    LEFT(rp.register_date, 7) AS reg_month,
                    COUNT(DISTINCT rp.id)     AS post_count
                FROM  recruit_posts             rp
                JOIN  recruit_post_specialties  rps ON rps.post_id = rp.id
                WHERE rp.register_date IS NOT NULL
                  AND rp.register_date <> ''
                  AND rp.region_sido   IS NOT NULL
                  AND rp.region_sido   <> ''
                GROUP BY rp.region_sido, rps.specialty, rp.employment_type,
                         LEFT(rp.register_date, 7)
                UNION ALL
                SELECT
                    (rp.region_sido || REGEXP_REPLACE(
                        SPLIT_PART(rp.region, ' ', 2), '(시|군)$', ''
                    ))                        AS region,
                    rps.specialty             AS specialty,
                    rp.employment_type        AS employment_type,
                    LEFT(rp.register_date, 7) AS reg_month,
                    COUNT(DISTINCT rp.id)     AS post_count
                FROM  recruit_posts             rp
                JOIN  recruit_post_specialties  rps ON rps.post_id = rp.id
                WHERE rp.register_date IS NOT NULL
                  AND rp.register_date <> ''
                  AND rp.region        IS NOT NULL
                  AND rp.region        <> ''
                  AND SPLIT_PART(rp.region, ' ', 2) ~ '(시|군)$'
                GROUP BY (rp.region_sido || REGEXP_REPLACE(
                              SPLIT_PART(rp.region, ' ', 2), '(시|군)$', ''
                          )),
                         rps.specialty, rp.employment_type,
                         LEFT(rp.register_date, 7)
                ORDER BY reg_month
            """), conn)
    except Exception as e:
        st.error(f"DB 연결 오류: {e}")
        return pd.DataFrame(columns=["region", "specialty", "employment_type",
                                     "reg_month", "post_count"])


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 병원 목록 조회 (클릭 시 호출 — 캐싱 없음, 매번 최신 조회)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def load_hospitals(month: str, region: str, specialty: str,
                   employment_type: str) -> pd.DataFrame:
    """
    선택된 월·지역·진료과·고용형태 조건에 해당하는 병원 목록을 반환.
    각 병원의 진료과가 여러 개인 경우 콤마로 합쳐서 1행으로 표시.
    """
    conditions = [
        "LEFT(rp.register_date, 7) = :month",
        "rp.register_date IS NOT NULL",
        "rp.register_date <> ''",
    ]
    params: dict = {"month": month}

    if region != "전체":
        if len(region) > 2:  # 시도+시군 조합 (예: 경기수원, 경북포항)
            sido = region[:2]
            city = region[2:]
            conditions.append("rp.region_sido = :region_sido")
            conditions.append("rp.region LIKE :region_like")
            params["region_sido"] = sido
            params["region_like"] = f"{sido} {city}%"
        else:
            conditions.append("rp.region_sido = :region")
            params["region"] = region
    if specialty != "전체":
        conditions.append("rps.specialty = :specialty")
        params["specialty"] = specialty
    if employment_type != "전체":
        conditions.append("rp.employment_type = :employment_type")
        params["employment_type"] = employment_type

    where = " AND ".join(conditions)

    # 중복횟수: 동일 진료과 기준으로 카운트
    # · specialty 필터가 있으면 해당 과 포함 공고만 카운트
    # · 전체면 현재 공고와 진료과가 하나라도 겹치는 공고만 카운트
    if specialty != "전체":
        count_subq = """(
            SELECT COUNT(DISTINCT rp2.id)
            FROM  recruit_posts rp2
            JOIN  recruit_post_specialties rps2 ON rps2.post_id = rp2.id
            WHERE rp2.hospital_name   = rp.hospital_name
              AND rp2.region_sido     = rp.region_sido
              AND rp2.employment_type = rp.employment_type
              AND rps2.specialty      = :specialty
        )"""
    else:
        count_subq = """(
            SELECT COUNT(DISTINCT rp2.id)
            FROM  recruit_posts rp2
            JOIN  recruit_post_specialties rps2 ON rps2.post_id = rp2.id
            WHERE rp2.hospital_name   = rp.hospital_name
              AND rp2.region_sido     = rp.region_sido
              AND rp2.employment_type = rp.employment_type
              AND rps2.specialty IN (
                  SELECT specialty FROM recruit_post_specialties
                  WHERE post_id = rp.id
              )
        )"""

    sql = text(f"""
        SELECT
            rp.hospital_name                        AS 병원명,
            CASE
                WHEN rp.region IS NOT NULL AND rp.region <> ''
                     AND SPLIT_PART(rp.region, ' ', 2) ~ '(시|군)$'
                THEN rp.region_sido || REGEXP_REPLACE(
                         SPLIT_PART(rp.region, ' ', 2), '(시|군)$', ''
                     )
                ELSE rp.region_sido
            END                                     AS 지역,
            rp.employment_type                      AS 고용형태,
            STRING_AGG(DISTINCT rps.specialty, ', '
                       ORDER BY rps.specialty)      AS 진료과,
            rp.salary_raw                           AS salary_raw,
            rp.salary_net_min                       AS salary_net_min,
            rp.salary_net_max                       AS salary_net_max,
            rp.register_date                        AS 등록일,
            rp.url                                  AS 공고링크,
            {count_subq}                            AS recruit_count
        FROM  recruit_posts             rp
        JOIN  recruit_post_specialties  rps ON rps.post_id = rp.id
        WHERE {where}
        GROUP BY rp.id, rp.hospital_name, rp.region_sido, rp.region,
                 rp.employment_type,
                 rp.salary_raw, rp.salary_net_min, rp.salary_net_max,
                 rp.register_date, rp.url
        ORDER BY rp.hospital_name
    """)

    try:
        with get_engine().connect() as conn:
            return pd.read_sql(sql, conn, params=params)
    except Exception as e:
        st.error(f"병원 목록 조회 오류: {e}")
        return pd.DataFrame()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 병원 구인 이력 조회
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def load_hospital_history(hospital_name: str, region_sido: str,
                          employment_type: str,
                          specialty: str = "전체") -> pd.DataFrame:
    """특정 병원의 구인 이력 (등록월·진료과·급여). specialty 필터 반영."""
    params = {
        "hospital_name":   hospital_name,
        "region_sido":     region_sido,
        "employment_type": employment_type,
    }
    # specialty 필터: 해당 과가 포함된 공고만 (전체면 전부 포함)
    if specialty != "전체":
        specialty_cond = """AND EXISTS (
            SELECT 1 FROM recruit_post_specialties s2
            WHERE s2.post_id = rp.id AND s2.specialty = :specialty
        )"""
        params["specialty"] = specialty
    else:
        specialty_cond = ""

    sql = text(f"""
        SELECT
            LEFT(rp.register_date, 7)               AS 등록월,
            (SELECT STRING_AGG(s.specialty, ', ' ORDER BY s.specialty)
             FROM   recruit_post_specialties s
             WHERE  s.post_id = rp.id)              AS 진료과,
            rp.salary_raw                           AS salary_raw,
            rp.salary_net_min                       AS salary_net_min,
            rp.salary_net_max                       AS salary_net_max,
            rp.url                                  AS 공고링크
        FROM  recruit_posts rp
        WHERE rp.hospital_name   = :hospital_name
          AND rp.region_sido     = :region_sido
          AND rp.employment_type = :employment_type
          {specialty_cond}
        ORDER BY rp.register_date
    """)
    try:
        with get_engine().connect() as conn:
            return pd.read_sql(sql, conn, params=params)
    except Exception as e:
        st.error(f"구인 이력 조회 오류: {e}")
        return pd.DataFrame()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 팝업 다이얼로그 — 병원 목록
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
@st.dialog("🏥 병원 목록", width="large")
def show_hospital_dialog(month: str, region: str, specialty: str,
                         employment_type: str):
    """막대 클릭 시 열리는 모달: 해당 월·조건의 병원 리스트."""

    # 헤더 정보
    region_label    = region          if region          != "전체" else "전체 지역"
    specialty_label = specialty       if specialty       != "전체" else "전체 진료과"
    emp_label       = employment_type if employment_type != "전체" else "전체 고용형태"
    st.markdown(
        f"**{month}** · {region_label} · {specialty_label} · {emp_label}",
        help="현재 적용된 필터 조건이 그대로 반영됩니다.",
    )
    st.divider()

    df_h = load_hospitals(month, region, specialty, employment_type)

    if df_h.empty:
        st.info("해당 조건의 병원 데이터가 없습니다.")
        return

    # ── 포맷 헬퍼 ──────────────────────────────────────────────────────────
    def format_salary(row):
        mn = row.get("salary_net_min")
        mx = row.get("salary_net_max")
        raw = row.get("salary_raw") or ""
        if mn is None or (isinstance(mn, float) and pd.isna(mn)):
            return raw[:20] + "…" if len(raw) > 20 else (raw or "-")
        mn, mx = int(mn), int(mx)
        if mn == mx:
            return f"{mn:,}만원"
        return f"{mn:,}~{mx:,}만원"

    def format_count(n):
        n = int(n)
        if n > 1:
            return f'<b style="color:#d32f2f">{n}회</b>'
        return f"{n}회"

    def make_link(url):
        if url and str(url).startswith("http"):
            return f'<a href="{url}" target="_blank">🔗 보기</a>'
        return "-"

    # ── 구인 이력 조회 (상단) ──────────────────────────────────────────────
    repeat_df = (
        df_h[df_h["recruit_count"] > 1][["병원명", "지역", "고용형태"]]
        .drop_duplicates()
        .reset_index(drop=True)
    )
    if not repeat_df.empty:
        st.markdown("#### 🔍 구인 이력 조회")
        st.caption(
            f"2회 이상 구인 공고를 올린 병원 **{len(repeat_df)}곳** — "
            "아래에서 병원을 선택하면 전체 이력을 확인할 수 있습니다."
        )
        options_map = {
            f"{r['병원명']}  ({r['지역']} / {r['고용형태']})": (
                r["병원명"], r["지역"], r["고용형태"]
            )
            for _, r in repeat_df.iterrows()
        }
        sel = st.selectbox(
            "병원 선택", ["─ 선택하세요 ─"] + list(options_map.keys()),
            key="hosp_hist_sel",
        )
        if sel != "─ 선택하세요 ─":
            h_name, h_region, h_emp = options_map[sel]
            df_hist = load_hospital_history(h_name, h_region, h_emp, specialty)
            if df_hist.empty:
                st.info("이력 데이터가 없습니다.")
            else:
                st.caption(
                    f"**{h_name}** ({h_region} / {h_emp})  —  총 {len(df_hist)}건"
                )
                df_hist.insert(
                    2, "Net월급(퇴직금포함)",
                    df_hist.apply(format_salary, axis=1),
                )
                df_hist["공고링크"] = df_hist["공고링크"].apply(make_link)
                df_hist = df_hist.drop(
                    columns=["salary_raw", "salary_net_min", "salary_net_max"]
                )
                st.markdown(
                    df_hist.to_html(escape=False, index=False),
                    unsafe_allow_html=True,
                )
        st.divider()

    # ── 병원 목록 표 ───────────────────────────────────────────────────────
    st.caption(f"총 **{len(df_h)}개** 병원")
    display = df_h.copy()
    display.insert(4, "Net월급(퇴직금포함)", display.apply(format_salary, axis=1))
    display.insert(6, "중복횟수", display["recruit_count"].apply(format_count))
    display["공고링크"] = display["공고링크"].apply(make_link)
    display = display.drop(
        columns=["salary_raw", "salary_net_min", "salary_net_max", "recruit_count"]
    )
    st.markdown(
        display.to_html(escape=False, index=False),
        unsafe_allow_html=True,
    )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 사이드바 — 필터
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with st.sidebar:
    st.header("🔍 필터")

    df_all = load_aggregated()

    if df_all.empty:
        st.warning("데이터가 없거나 DB 연결에 실패했습니다.")
        st.stop()

    region_options  = ["전체"] + sorted(df_all["region"].dropna().unique().tolist())
    selected_region = st.selectbox("📍 지역", region_options)

    if selected_region == "전체":
        specialty_pool = sorted(df_all["specialty"].dropna().unique().tolist())
    else:
        specialty_pool = sorted(
            df_all[df_all["region"] == selected_region]["specialty"]
            .dropna().unique().tolist()
        )
    selected_specialty = st.selectbox("🩺 진료과", ["전체"] + specialty_pool)

    st.divider()
    if st.button("🔄 데이터 새로고침"):
        st.cache_data.clear()
        st.rerun()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 필터링 & 월별 집계  (employment_type 필터는 차트 섹션에서 선택 후 적용)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
df = df_all.copy()
if selected_region    != "전체":
    df = df[df["region"]    == selected_region]
if selected_specialty != "전체":
    df = df[df["specialty"] == selected_specialty]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 메인 화면
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.title("🏥 구인 트렌드 대시보드")
st.caption(f"필터 적용 중 → 지역: **{selected_region}** · 진료과: **{selected_specialty}**")
st.divider()

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 막대그래프 — 제목 + 고용형태 드롭다운 (인라인)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 고용형태 목록: DB 실제 값 기준 (건수 많은 순 고정 정렬)
EMPLOYMENT_TYPES = [
    "전체", "봉직의", "대진의", "당직의", "전임의", "전공의",
    "입원전담전문의", "출장검진", "임상(사내의사)", "임상외", "동업", "기타",
]

col_title, col_emp = st.columns([5, 2])
with col_title:
    st.subheader("📊 월별 구인건수")
    st.caption("💡 막대를 클릭하면 해당 월의 병원 목록을 확인할 수 있습니다.")
with col_emp:
    st.markdown("<br>", unsafe_allow_html=True)   # 제목과 높이 맞춤
    selected_employment = st.selectbox(
        "👔 고용형태",
        EMPLOYMENT_TYPES,
        index=0,                   # 기본값: 전체
        key="employment_filter",
    )

# 고용형태 필터 적용 후 월별 집계
df_emp = df.copy()
if selected_employment != "전체":
    df_emp = df_emp[df_emp["employment_type"] == selected_employment]

df_monthly = (
    df_emp.groupby("reg_month")["post_count"]
    .sum().reset_index().sort_values("reg_month")
)
df_monthly["reg_month"] = df_monthly["reg_month"].astype(str)

# ── KPI 카드 ───────────────────────────────────────────────────────────────
col1, col2, col3 = st.columns(3)
total_posts  = int(df_monthly["post_count"].sum()) if not df_monthly.empty else 0
col1.metric("총 공고 수",  f"{total_posts:,}건")
col2.metric("집계 월 수",  f"{len(df_monthly)}개월")
if not df_monthly.empty:
    peak = df_monthly.loc[df_monthly["post_count"].idxmax()]
    col3.metric("최고 공고월", f"{peak['reg_month']} ({int(peak['post_count'])}건)")
else:
    col3.metric("최고 공고월", "-")

st.divider()

if df_monthly.empty:
    st.warning("선택한 조건에 해당하는 데이터가 없습니다.")
else:
    fig = px.bar(
        df_monthly,
        x="reg_month",
        y="post_count",
        text="post_count",
        labels={"reg_month": "등록 월", "post_count": "공고 수"},
        color_discrete_sequence=["#2196F3"],
        category_orders={"reg_month": sorted(df_monthly["reg_month"].tolist())},
    )
    fig.update_traces(
        textposition="outside",
        textfont_size=12,
        hovertemplate="<b>%{x}</b><br>공고 수: <b>%{y}건</b>  ← 클릭하세요<extra></extra>",
    )
    fig.update_layout(
        xaxis_title="등록 월",
        yaxis_title="공고 수",
        plot_bgcolor="white",
        xaxis=dict(tickangle=-30, type="category"),
        yaxis=dict(
            gridcolor="#eeeeee",
            zeroline=True,
            range=[0, df_monthly["post_count"].max() * 1.25],
        ),
        bargap=0.35,
        height=460,
        margin=dict(t=30, b=50, l=50, r=20),
        hoverlabel=dict(bgcolor="white", font_size=13),
        # 클릭 가능함을 커서로 암시
        clickmode="event",
    )

    # on_select="rerun": 클릭 시 앱 재실행 + 선택 정보 반환
    event = st.plotly_chart(
        fig,
        use_container_width=True,
        on_select="rerun",
        selection_mode="points",
        key="bar_chart",
    )

    # 클릭된 막대가 있으면 다이얼로그 열기
    points = event.selection.get("points", []) if event.selection else []
    if points:
        clicked_month = str(points[0].get("x", ""))
        if clicked_month:
            show_hospital_dialog(
                clicked_month, selected_region,
                selected_specialty, selected_employment,
            )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 상세 데이터 테이블
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with st.expander("📋 상세 데이터 테이블 보기"):
    if df_emp.empty:
        st.info("데이터가 없습니다.")
    else:
        st.dataframe(
            df_emp.sort_values(["reg_month", "region", "specialty"])
            .rename(columns={
                "region":          "지역",
                "specialty":       "진료과",
                "employment_type": "고용형태",
                "reg_month":       "등록 월",
                "post_count":      "공고 수",
            })
            .reset_index(drop=True),
            use_container_width=True,
            hide_index=True,
        )


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 급여 현황 — 월별 Net 월급 추이
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
st.divider()
st.subheader("💰 급여 현황 — 월별 평균 Net 월급 추이 (봉직의)")
st.caption(
    "봉직의 공고 한정 · Net 월급 기준 · 인센티브 비포함 · 협의/미기재 공고 제외 · "
    "15건 이상 그룹: IQR 이상치 제거 후 평균 · 15건 미만 그룹: 중앙값"
)

# ── 데이터 로드 (사이드바 필터 그대로 사용) ───────────────────────────────
df_sal = load_salary_monthly(selected_region, selected_specialty)

# ── KPI 카드 ───────────────────────────────────────────────────────────────
sk1, sk2, sk3 = st.columns(3)

if not df_sal.empty:
    total_cnt = int(df_sal["공고수"].sum())
    overall_avg = int(
        (df_sal["평균Net월급"] * df_sal["공고수"]).sum() / df_sal["공고수"].sum()
    )
    peak = df_sal.loc[df_sal["평균Net월급"].idxmax()]
    sk1.metric("집계 공고 수", f"{total_cnt:,}건")
    sk2.metric("전체 기간 평균", f"{overall_avg:,}만원")
    sk3.metric("최고 평균 월", f"{peak['등록월']} ({int(peak['평균Net월급']):,}만원)")
else:
    sk1.metric("집계 공고 수", "-")
    sk2.metric("전체 기간 평균", "-")
    sk3.metric("최고 평균 월", "-")

st.divider()

# ── 월별 막대 그래프 ────────────────────────────────────────────────────────
if df_sal.empty:
    st.info("선택한 조건에 해당하는 급여 데이터가 없습니다.")
else:
    region_label    = selected_region    if selected_region    != "전체" else "전국"
    specialty_label = selected_specialty if selected_specialty != "전체" else "전체 진료과"
    chart_title = f"{region_label} · {specialty_label} 월별 평균 Net 월급"

    fig_sal = px.bar(
        df_sal,
        x="등록월",
        y="평균Net월급",
        text=df_sal["평균Net월급"].apply(lambda v: f"{int(v):,}만원"),
        custom_data=["공고수"],
        color_discrete_sequence=["#43A047"],
        category_orders={"등록월": sorted(df_sal["등록월"].tolist())},
        title=chart_title,
    )
    fig_sal.update_traces(
        textposition="outside",
        textfont_size=12,
        hovertemplate=(
            "<b>%{x}</b><br>"
            "평균 Net 월급: <b>%{y:,}만원</b><br>"
            "집계 공고: %{customdata[0]}건<extra></extra>"
        ),
    )
    fig_sal.update_layout(
        xaxis_title="등록 월",
        yaxis_title="평균 Net 월급 (만원)",
        plot_bgcolor="white",
        xaxis=dict(tickangle=-30, type="category"),
        yaxis=dict(
            gridcolor="#eeeeee",
            zeroline=True,
            range=[
                max(0, df_sal["평균Net월급"].min() * 0.85),
                df_sal["평균Net월급"].max() * 1.15,
            ],
        ),
        bargap=0.35,
        height=460,
        margin=dict(t=50, b=50, l=60, r=20),
        hoverlabel=dict(bgcolor="white", font_size=13),
        title_font_size=15,
    )
    st.plotly_chart(fig_sal, use_container_width=True)

# ── 지역별 / 진료과별 전체 순위 (참고용) ──────────────────────────────────
_spec_label = selected_specialty if selected_specialty != "전체" else "전체 진료과"
_expander_title = (
    f"📊 지역별 · 진료과별 평균 순위 보기  |  진료과: {_spec_label}"
    if selected_specialty != "전체"
    else "📊 지역별 · 진료과별 평균 순위 보기"
)

with st.expander(_expander_title):
    df_rank_r, df_rank_s = load_salary_ranking(selected_region, selected_specialty)
    tab_r, tab_s = st.tabs(["📍 지역별", "🩺 진료과별"])

    with tab_r:
        if df_rank_r.empty:
            st.info("데이터가 없습니다.")
        else:
            st.caption(
                f"진료과 기준: **{_spec_label}** "
                f"{'· 해당 진료과 공고만 집계' if selected_specialty != '전체' else '· 전체 진료과 공고 집계'}"
            )
            df_plot_r = df_rank_r.head(17).sort_values("평균Net월급")
            fig_r = px.bar(
                df_plot_r,
                x="평균Net월급", y="지역", orientation="h",
                text=df_plot_r["평균Net월급"].apply(lambda v: f"{int(v):,}만원"),
                custom_data=["공고수"],
                color="평균Net월급", color_continuous_scale="Blues",
                title=f"지역별 평균 Net 월급 ({_spec_label})",
            )
            fig_r.update_traces(
                textposition="outside", textfont_size=11,
                hovertemplate=(
                    "<b>%{y}</b><br>평균 Net 월급: <b>%{x:,}만원</b><br>"
                    "집계 공고: %{customdata[0]}건<extra></extra>"
                ),
            )
            fig_r.update_layout(
                xaxis_title="평균 Net 월급 (만원)", yaxis_title="",
                plot_bgcolor="white", xaxis=dict(gridcolor="#eeeeee"),
                coloraxis_showscale=False,
                height=max(300, len(df_plot_r) * 36),
                margin=dict(t=40, b=40, l=10, r=80),
                title_font_size=14,
            )
            st.plotly_chart(fig_r, use_container_width=True)

    with tab_s:
        if df_rank_s.empty:
            st.info("데이터가 없습니다.")
        else:
            df_plot_s = df_rank_s.head(20).sort_values("평균Net월급")
            fig_s = px.bar(
                df_plot_s,
                x="평균Net월급", y="진료과", orientation="h",
                text=df_plot_s["평균Net월급"].apply(lambda v: f"{int(v):,}만원"),
                custom_data=["공고수"],
                color="평균Net월급", color_continuous_scale="Greens",
            )
            fig_s.update_traces(
                textposition="outside", textfont_size=11,
                hovertemplate=(
                    "<b>%{y}</b><br>평균 Net 월급: <b>%{x:,}만원</b><br>"
                    "집계 공고: %{customdata[0]}건<extra></extra>"
                ),
            )
            fig_s.update_layout(
                xaxis_title="평균 Net 월급 (만원)", yaxis_title="",
                plot_bgcolor="white", xaxis=dict(gridcolor="#eeeeee"),
                coloraxis_showscale=False,
                height=max(300, len(df_plot_s) * 36),
                margin=dict(t=10, b=40, l=10, r=80),
            )
            st.plotly_chart(fig_s, use_container_width=True)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 마취통증의학과 장기 트렌드 — 엑셀(과거) + DB(크롤링) 통합
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
if selected_specialty == "마취통증의학과":
    st.divider()
    st.subheader("💉 마취통증의학과 장기 트렌드 (엑셀 과거자료 + DB 통합)")
    st.caption(
        "엑셀: 2023-03 ~ 2026-01 (수동 수집 · Net 월급 기준) │ "
        "DB: 크롤링 데이터 (net/monthly 공고만 급여 집계) │ "
        "2026-01은 양쪽 출처 모두 표시"
    )

    df_xls = load_excel_machwi()
    df_dbc = load_db_machwi()

    if df_xls.empty and df_dbc.empty:
        st.warning("마취통증의학과 데이터를 불러올 수 없습니다.")
    else:
        # ── KPI 카드 ───────────────────────────────────────────────────────
        kc1, kc2, kc3, kc4 = st.columns(4)
        kc1.metric("엑셀 수집 개월수", f"{len(df_xls)}개월" if not df_xls.empty else "-")
        kc2.metric("DB 수집 개월수",   f"{len(df_dbc)}개월" if not df_dbc.empty else "-")
        if not df_xls.empty:
            kc3.metric(
                "엑셀 전체 평균 Net 월급",
                f"{int((df_xls['평균Net월급'] * df_xls['공고수']).sum() / df_xls['공고수'].sum()):,}만원",
            )
        else:
            kc3.metric("엑셀 전체 평균 Net 월급", "-")
        if not df_dbc.empty and df_dbc["평균Net월급"].notna().any():
            kc4.metric(
                "DB 전체 평균 Net 월급",
                f"{int(df_dbc['평균Net월급'].dropna().mean()):,}만원",
            )
        else:
            kc4.metric("DB 전체 평균 Net 월급", "-")

        st.divider()

        COLOR_XLS = "#2196F3"
        COLOR_DBC = "#FF9800"

        # ── 차트 1: 월별 구인 공고수 ──────────────────────────────────────
        st.markdown("#### 📊 월별 구인 공고수")
        fig_cnt = go.Figure()
        if not df_xls.empty:
            fig_cnt.add_trace(go.Bar(
                x=df_xls["등록월"],
                y=df_xls["공고수"],
                name="엑셀(과거)",
                marker_color=COLOR_XLS,
                text=df_xls["공고수"],
                textposition="outside",
                hovertemplate="<b>%{x}</b><br>공고수: <b>%{y}건</b> [엑셀]<extra></extra>",
            ))
        if not df_dbc.empty:
            fig_cnt.add_trace(go.Bar(
                x=df_dbc["등록월"],
                y=df_dbc["공고수"],
                name="DB(크롤링)",
                marker_color=COLOR_DBC,
                text=df_dbc["공고수"],
                textposition="outside",
                hovertemplate="<b>%{x}</b><br>공고수: <b>%{y}건</b> [DB]<extra></extra>",
            ))
        fig_cnt.update_layout(
            barmode="group",
            xaxis=dict(title="등록 월", tickangle=-30, type="category",
                       categoryorder="category ascending"),
            yaxis=dict(title="공고 수", gridcolor="#eeeeee", zeroline=True),
            plot_bgcolor="white",
            bargap=0.25, bargroupgap=0.1, height=420,
            margin=dict(t=20, b=60, l=50, r=20),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            hoverlabel=dict(bgcolor="white", font_size=13),
        )
        st.plotly_chart(fig_cnt, use_container_width=True)

        # ── 차트 2: 월별 평균 Net 월급 ────────────────────────────────────
        st.markdown("#### 💰 월별 평균 Net 월급 추이")
        st.caption("엑셀: 공고 기재 급여 평균 │ DB: net/monthly 공고만 집계 (협의·미기재 제외)")
        fig_sal = go.Figure()
        if not df_xls.empty:
            fig_sal.add_trace(go.Scatter(
                x=df_xls["등록월"],
                y=df_xls["평균Net월급"],
                mode="lines+markers+text",
                name="엑셀(과거)",
                line=dict(color=COLOR_XLS, width=2),
                marker=dict(size=7),
                text=df_xls["평균Net월급"].apply(lambda v: f"{int(v):,}"),
                textposition="top center",
                textfont=dict(size=10, color=COLOR_XLS),
                hovertemplate=(
                    "<b>%{x}</b><br>평균 Net 월급: <b>%{y:,}만원</b> [엑셀]<extra></extra>"
                ),
            ))
        if not df_dbc.empty:
            df_sal_dbc = df_dbc.dropna(subset=["평균Net월급"])
            if not df_sal_dbc.empty:
                fig_sal.add_trace(go.Scatter(
                    x=df_sal_dbc["등록월"],
                    y=df_sal_dbc["평균Net월급"],
                    mode="lines+markers+text",
                    name="DB(크롤링)",
                    line=dict(color=COLOR_DBC, width=2, dash="dash"),
                    marker=dict(size=9, symbol="diamond"),
                    text=df_sal_dbc["평균Net월급"].apply(lambda v: f"{int(v):,}"),
                    textposition="top center",
                    textfont=dict(size=10, color=COLOR_DBC),
                    hovertemplate=(
                        "<b>%{x}</b><br>평균 Net 월급: <b>%{y:,}만원</b> [DB]<extra></extra>"
                    ),
                ))
        all_vals = []
        if not df_xls.empty:
            all_vals += df_xls["평균Net월급"].tolist()
        if not df_dbc.empty:
            all_vals += df_dbc["평균Net월급"].dropna().tolist()
        y_min = max(0, min(all_vals) * 0.90) if all_vals else 0
        y_max = max(all_vals) * 1.12         if all_vals else 5000
        fig_sal.update_layout(
            xaxis=dict(title="등록 월", tickangle=-30, type="category",
                       categoryorder="category ascending"),
            yaxis=dict(title="평균 Net 월급 (만원)", gridcolor="#eeeeee",
                       zeroline=False, range=[y_min, y_max]),
            plot_bgcolor="white", height=440,
            margin=dict(t=20, b=60, l=60, r=20),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            hoverlabel=dict(bgcolor="white", font_size=13),
        )
        st.plotly_chart(fig_sal, use_container_width=True)
