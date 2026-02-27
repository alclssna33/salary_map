#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_excel_to_db.py  —  (마봉협)구인구직정리.xlsx → machwi_excel_history 테이블 import

실행
----
    python import_excel_to_db.py           # 기본 실행 (중복 방지 후 insert)
    python import_excel_to_db.py --reset   # 기존 데이터 전부 삭제 후 재import

구조
----
    엑셀 '일자리분석' 시트: 4열 1그룹 (blank | 지역 | 병원명 | Pay) × 35개월
    Pay 단위: 2.3 → 2,300만원 (net_pay = 2300)
    source = 'excel_import' 으로 저장 (수동 수집 원본임을 명시)
"""

import argparse
import os
import re
import sys
from datetime import datetime

import openpyxl
import psycopg2

# ─────────────────────────────────────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────────────────────────────────────
DB_CONFIG = {
    "host":     "localhost",
    "port":     5432,
    "dbname":   "medigate",
    "user":     "postgres",
    "password": "postgres",
}

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "(마봉협)구인구직정리.xlsx",
)

SOURCE_TAG = "excel_import"   # DB에 남길 출처 표시


# ─────────────────────────────────────────────────────────────────────────────
# 엑셀 파싱
# ─────────────────────────────────────────────────────────────────────────────
def parse_excel() -> list[dict]:
    """일자리분석 시트에서 병원별 raw 행 추출."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["일자리분석"]

    records = []
    col = 1
    while col <= ws.max_column:
        # ── 날짜 파싱 (row 8, 그룹 첫 번째 열) ──────────────────────────────
        date_val = ws.cell(8, col).value
        if date_val is None:
            col += 4
            continue

        if isinstance(date_val, datetime):
            month_str = date_val.strftime("%Y-%m")
        elif isinstance(date_val, str):
            m = re.match(r"(\d{4})년(\d{1,2})월", date_val)
            month_str = f"{m.group(1)}-{int(m.group(2)):02d}" if m else None
        else:
            month_str = None

        if not month_str:
            col += 4
            continue

        # ── 데이터 행 수집 (row 10~) ─────────────────────────────────────────
        # 중간에 빈 행(gap)이 있을 수 있으므로 break 없이 끝까지 순회
        for row in range(10, ws.max_row + 1):
            region   = ws.cell(row, col + 1).value
            hospital = ws.cell(row, col + 2).value
            pay_raw  = ws.cell(row, col + 3).value

            # Pay 유효성 검사: 0.5 ~ 10.0 범위 (500 ~ 10,000만원)
            try:
                pay_f = float(pay_raw)
                if not (0.5 <= pay_f <= 10.0):
                    continue
                net_pay = round(pay_f * 1000)
            except (TypeError, ValueError):
                continue

            records.append({
                "reg_month":     month_str,
                "region":        str(region).strip() if region else None,
                "hospital_name": str(hospital).strip() if hospital else None,
                "net_pay":       net_pay,
            })

        col += 4

    return records


# ─────────────────────────────────────────────────────────────────────────────
# DB insert
# ─────────────────────────────────────────────────────────────────────────────
def import_to_db(records: list[dict], reset: bool = False):
    conn = psycopg2.connect(**DB_CONFIG)
    cur  = conn.cursor()

    if reset:
        cur.execute("DELETE FROM machwi_excel_history WHERE source = %s", (SOURCE_TAG,))
        print(f"  기존 {SOURCE_TAG} 데이터 삭제 완료")

    # 기존 (reg_month, hospital_name, region) 중복 방지
    cur.execute("""
        SELECT reg_month, hospital_name, region
        FROM   machwi_excel_history
        WHERE  source = %s
    """, (SOURCE_TAG,))
    existing = {(r[0], r[1], r[2]) for r in cur.fetchall()}

    inserted = skipped = 0
    for r in records:
        key = (r["reg_month"], r["hospital_name"], r["region"])
        if key in existing:
            skipped += 1
            continue
        cur.execute("""
            INSERT INTO machwi_excel_history
                (reg_month, region, hospital_name, net_pay, source)
            VALUES (%s, %s, %s, %s, %s)
        """, (r["reg_month"], r["region"], r["hospital_name"], r["net_pay"], SOURCE_TAG))
        existing.add(key)
        inserted += 1

    conn.commit()
    conn.close()
    return inserted, skipped


# ─────────────────────────────────────────────────────────────────────────────
# 결과 요약 출력
# ─────────────────────────────────────────────────────────────────────────────
def print_summary(conn_cfg):
    conn = psycopg2.connect(**conn_cfg)
    cur  = conn.cursor()
    cur.execute("""
        SELECT reg_month, COUNT(*) AS cnt, ROUND(AVG(net_pay)) AS avg_pay
        FROM   machwi_excel_history
        WHERE  source = %s
        GROUP  BY reg_month
        ORDER  BY reg_month
    """, (SOURCE_TAG,))
    rows = cur.fetchall()
    conn.close()

    print(f"\n{'월':^10} {'공고수':^6} {'평균Net월급':^12}")
    print("-" * 32)
    for reg_month, cnt, avg_pay in rows:
        print(f"  {reg_month}   {cnt:>4}건   {int(avg_pay):>6,}만원")
    print(f"\n  총 {sum(r[1] for r in rows):,}건 / {len(rows)}개월")


# ─────────────────────────────────────────────────────────────────────────────
# main
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="엑셀 → machwi_excel_history import")
    parser.add_argument("--reset", action="store_true",
                        help="기존 excel_import 데이터 삭제 후 재import")
    args = parser.parse_args()

    if not os.path.exists(EXCEL_PATH):
        print(f"[오류] 엑셀 파일을 찾을 수 없습니다: {EXCEL_PATH}")
        sys.exit(1)

    print("엑셀 파싱 중...")
    records = parse_excel()
    print(f"  파싱 완료: {len(records):,}행 ({len({r['reg_month'] for r in records})}개월)")

    print(f"\nDB import 중... (reset={args.reset})")
    inserted, skipped = import_to_db(records, reset=args.reset)
    print(f"  INSERT {inserted:,}건 / 중복 스킵 {skipped:,}건")

    print("\n─── 월별 import 결과 ───")
    print_summary(DB_CONFIG)
    print("\n완료.")


if __name__ == "__main__":
    main()
